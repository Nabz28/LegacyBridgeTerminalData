import openpyxl, json, re, os, glob
ROOT = r"C:\Users\DELL\Downloads\Claude Repository\Projects\Macro Terminal v1\CEIC Data\02_Industry"
OUT  = os.path.dirname(os.path.abspath(__file__))

FREQ = {'daily': 'P1D', 'weekly': 'P7D', 'monthly': 'P1M', 'quarterly': 'P3M',
        'annual': 'P1Y', 'yearly': 'P1Y', 'half': 'P6M', 'semi': 'P6M'}


def freq_iso(s):
    s = (s or '').lower()
    for k, v in FREQ.items():
        if k in s:
            return v
    return 'NA'


def hdr_is_date(c):
    return hasattr(c, 'year')


def parse_file(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    date_start = None
    for i, c in enumerate(header):
        if hdr_is_date(c):
            date_start = i
            break
    if date_start is None:
        date_start = 27
    dates = [(header[j].date().isoformat() if hdr_is_date(header[j]) else None)
             for j in range(date_start, len(header))]
    series = []
    for r in rows:
        if r is None:
            continue
        name = r[0]
        if name is None or str(name).strip() == '':
            continue
        sid_raw = r[7] if len(r) > 7 else None
        m = re.match(r'\s*(\d+)\s*(?:\(([^)]*)\))?', str(sid_raw or ''))
        numid = m.group(1) if m else None
        mnem = (m.group(2) if m and m.group(2) else (r[10] if len(r) > 10 else None))
        obs = []
        for k, j in enumerate(range(date_start, len(header))):
            v = r[j] if j < len(r) else None
            d = dates[k]
            if v is None or d is None:
                continue
            if isinstance(v, (int, float)):
                obs.append([d, float(v)])
        status = str(r[6] if len(r) > 6 else '')
        series.append({
            'name': str(name).strip(),
            'region': r[1] if len(r) > 1 else None,
            'frequency_raw': r[3] if len(r) > 3 else None,
            'frequency': freq_iso(r[3] if len(r) > 3 else ''),
            'unit': r[4] if len(r) > 4 else None,
            'source': r[5] if len(r) > 5 else None,
            'status': status,
            'series_id': numid,
            'mnemonic': mnem,
            'first_obs': str(r[12]) if len(r) > 12 and r[12] else (obs[0][0] if obs else None),
            'last_obs': str(r[13]) if len(r) > 13 and r[13] else (obs[-1][0] if obs else None),
            'n_obs_meta': r[26] if len(r) > 26 else None,
            'n_obs': len(obs),
            'discontinued': bool(re.search(r'\(DC\)', str(name))) or status.lower().startswith('disc'),
            'obs': obs,
        })
    wb.close()
    return series


def main():
    files = sorted(glob.glob(os.path.join(ROOT, '**', '*.xlsx'), recursive=True))
    master = {}
    summ = []
    tot_series = tot_obs = 0
    for f in files:
        rel = os.path.relpath(f, ROOT).replace(os.sep, '/')
        try:
            s = parse_file(f)
        except Exception as e:
            summ.append((rel, 'ERR', str(e)[:80]))
            continue
        master[rel] = s
        no = sum(x['n_obs'] for x in s)
        tot_series += len(s)
        tot_obs += no
        summ.append((rel, len(s), no))
    with open(os.path.join(OUT, 'master.json'), 'w', encoding='utf-8') as fh:
        json.dump(master, fh)
    print('FILES', len(files), 'SERIES', tot_series, 'OBS', tot_obs)
    print('--- per file: n_series | n_obs | rel ---')
    for rel, a, b in summ:
        print('%5s %8s  %s' % (a, b, rel))


main()
