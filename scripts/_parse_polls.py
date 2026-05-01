"""Parser for Reuters/Refinitiv "Economic Indicator Polls" xlsx files.

Schema (verified consistent across all 42 polls files):
  R1C1 = 'Economic Indicator Polls'
  R2C1 = 'Indicator', R2C2 = 'United States - <Indicator Name>'
  R3C1 = 'Download date', R3C2 = <date string>
  R4C1 = 'Unit',          R4C2 = 'Index' / 'Percent' / etc.
  R5     = blank
  R6     = 'Period' | 'Consensus' | (cols 12+ 'Contributor')
  R7     = headers: blank | Actual | Median | SmartEconomics | Predicted Surprise | Mean | Mode | Min | Max | Standard Deviation | # Forecasts | <contributor names...>
  R8     = RIC row: blank | <ACTUAL_RIC> | p<CODE>=M | p<CODE>=F | p<CODE>=P | p<CODE>=E | p<CODE>=O | p<CODE>=L | p<CODE>=H | p<CODE>=T | p<CODE>=C | <contributor RICs...>
  R9+    = observations: <date> | actual | median | smartecon | pred_surprise | mean | mode | min | max | stddev | n_forecasts | <contributor values...>

This parser EMITS one Series per (indicator × statistic) combination:
  - The "Actual" RIC (USPMI=ECI etc.) — the realised release
  - 10 statistic RICs (Median, SmartEcon, Predicted Surprise, Mean, Mode, Min, Max, StdDev, # Forecasts)
  - SKIPS per-contributor RICs (low analytical value, hundreds of them)

All RICs tagged with section="Polls", category="Polls — <Indicator Name>",
category_slug="us_polls".
"""
from __future__ import annotations

import os
import re
from dataclasses import dataclass, field
from datetime import datetime
from typing import Iterator

import openpyxl


POLLS_DIR = r"C:\Users\DELL\Downloads\Refinitiv Integration\US Macro Data\US Polls"


@dataclass
class PollSeries:
    ric: str
    description: str          # e.g. "ISM Manufacturing PMI — Median Consensus"
    indicator: str            # e.g. "ISM Manufacturing PMI"
    statistic: str            # "Actual" | "Median" | "Mean" | ...
    unit: str
    frequency: str
    section: str = "Polls"
    category: str = "Polls"
    category_slug: str = "us_polls"
    source_file: str = ""
    observations: list[tuple[str, float]] = field(default_factory=list)


# Column positions (0-indexed) and statistic labels
STAT_COLS = [
    (1, "Actual"),
    (2, "Median Consensus"),
    (3, "SmartEconomics Forecast"),
    (4, "Predicted Surprise"),
    (5, "Mean Forecast"),
    (6, "Mode Forecast"),
    (7, "Min Forecast"),
    (8, "Max Forecast"),
    (9, "Forecast Standard Deviation"),
    (10, "Number of Forecasters"),
]


def list_polls_xlsx(polls_dir: str = POLLS_DIR) -> list[str]:
    files = []
    for name in sorted(os.listdir(polls_dir)):
        if name.lower().endswith(".xlsx") and not name.startswith("~$"):
            files.append(os.path.join(polls_dir, name))
    return files


def detect_freq(dates: list[datetime]) -> str:
    if len(dates) < 2:
        return ""
    deltas = []
    for i in range(1, min(20, len(dates))):
        d = abs((dates[i - 1] - dates[i]).days)
        if d > 0:
            deltas.append(d)
    if not deltas:
        return ""
    import statistics as st
    median = st.median(deltas)
    if median <= 2: return "P1D"
    if median <= 10: return "P1W"
    if median <= 45: return "P1M"
    if median <= 120: return "P3M"
    if median <= 200: return "P6M"
    return "P1Y"


def parse_polls_workbook(path: str) -> Iterator[PollSeries]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb.active

    # R2C2: indicator name; strip "United States - " prefix
    indicator_full = (ws.cell(2, 2).value or "").strip()
    indicator = re.sub(r"^United States\s*-\s*", "", indicator_full).strip()
    if not indicator:
        # Fall back to filename
        indicator = os.path.splitext(os.path.basename(path))[0].replace("US_", "").replace("US ", "").strip()

    unit = (ws.cell(4, 2).value or "").strip()

    # Read RIC row (R8) and data rows (R9+)
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 9:
        wb.close()
        return

    ric_row = rows[7]  # R8 (0-indexed = 7)

    # Build per-statistic series
    series_by_col: dict[int, PollSeries] = {}
    for col_idx, stat_label in STAT_COLS:
        if col_idx >= len(ric_row):
            continue
        ric_val = ric_row[col_idx]
        if ric_val is None or str(ric_val).strip() == "":
            continue
        ric = str(ric_val).strip()
        desc = f"{indicator} — {stat_label}"
        series_by_col[col_idx] = PollSeries(
            ric=ric,
            description=desc,
            indicator=indicator,
            statistic=stat_label,
            unit=unit,
            frequency="",
            source_file=os.path.basename(path),
            observations=[],
        )

    # Walk data rows (R9+, idx 8+)
    date_objs: dict[int, list[datetime]] = {c: [] for c in series_by_col}
    for row in rows[8:]:
        if not row or row[0] is None:
            continue
        d = row[0]
        if isinstance(d, datetime):
            date_obj = d
        else:
            try:
                date_obj = datetime.fromisoformat(str(d))
            except (TypeError, ValueError):
                continue
        for col_idx, ser in series_by_col.items():
            if col_idx >= len(row):
                continue
            v = row[col_idx]
            if v is None:
                continue
            try:
                value = float(v)
            except (TypeError, ValueError):
                continue
            ser.observations.append((date_obj.date().isoformat(), value))
            date_objs[col_idx].append(date_obj)

    wb.close()

    # Sort + dedupe + detect frequency, then yield
    for col_idx, ser in series_by_col.items():
        seen: dict[str, float] = {}
        for d, v in ser.observations:
            seen[d] = v
        ser.observations = sorted(seen.items())
        ser.frequency = detect_freq(sorted(date_objs[col_idx], reverse=True))
        if ser.observations:  # skip empty series
            yield ser


def parse_all_polls(polls_dir: str = POLLS_DIR) -> Iterator[PollSeries]:
    for path in list_polls_xlsx(polls_dir):
        yield from parse_polls_workbook(path)


if __name__ == "__main__":
    # Quick self-test: list all RICs we'd produce
    total = 0
    by_indicator: dict[str, int] = {}
    for s in parse_all_polls():
        total += 1
        by_indicator[s.indicator] = by_indicator.get(s.indicator, 0) + 1
    print(f"[polls] {total} RICs across {len(by_indicator)} indicators")
    for ind, n in sorted(by_indicator.items()):
        print(f"  {ind:50s}  {n} RICs")
