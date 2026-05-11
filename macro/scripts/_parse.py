"""Shared parser for Refinitiv macro xlsx files (country-agnostic).

Schema (verified consistent across all US and Indonesia files):
  R1C1 = country header
  R4C1 = section ("Money & Finance", "Prices", ...)
  R4C2 = subcategory ("Banking", "Consumer Prices / Inflation", ...)
  R4 cols 4,7,10,...  = per-RIC description
  R5 cols 4,7,10,...  = RIC code
  R7+ alternating (date, value) at (col, col+1), spacer at col+2

Country selection (in priority order):
  1. RIC_COUNTRY env var ("us", "id", ...) — picks the dir + slug prefix
  2. RIC_XLSX_DIR env var — explicit override of the source directory
  3. Default: "us"
"""

from __future__ import annotations

import os
import re
import statistics
from dataclasses import dataclass, field
from datetime import datetime
from typing import Iterator

import openpyxl


# Per-country source directories. Add new countries here as needed.
COUNTRY_DIRS = {
    "us": r"C:\Users\DELL\Downloads\Refinitiv Integration\US Macro Data\US Macro",
    "id": r"C:\Users\DELL\Downloads\Refinitiv Integration\Indonesia Macro Data\Indonesia Macro",
    "cn": r"C:\Users\DELL\Downloads\Refinitiv Integration\China Macro Data",
}


def get_country() -> str:
    return os.environ.get("RIC_COUNTRY", "us").lower()


def get_xlsx_dir(country: str | None = None) -> str:
    explicit = os.environ.get("RIC_XLSX_DIR")
    if explicit:
        return explicit
    cc = (country or get_country()).lower()
    if cc not in COUNTRY_DIRS:
        raise ValueError(f"unknown country code: {cc!r} (known: {list(COUNTRY_DIRS)})")
    return COUNTRY_DIRS[cc]


# Backward-compat: many scripts import XLSX_DIR directly. Resolve at import time.
XLSX_DIR = get_xlsx_dir()


@dataclass
class Series:
    ric: str
    description: str
    frequency: str  # "P1D" | "P1W" | "P1M" | "P3M" | "P1Y" | ""
    section: str
    category: str
    category_slug: str
    source_file: str
    observations: list[tuple[str, float]] = field(default_factory=list)  # (iso_date, value)


def slugify(text: str) -> str:
    s = text.lower()
    s = re.sub(r"[^a-z0-9]+", "_", s)
    s = re.sub(r"_+", "_", s).strip("_")
    return s


def slug_from_filename(filename: str, country: str | None = None) -> str:
    """Build a category slug like 'us_banking' or 'id_banking' from a workbook
    filename. Strips any leading country prefix in the file name and prepends
    the canonical country code so per-country slugs never collide.

    Also handles the China filename convention which embeds the category in a
    trailing parenthetical with optional section prefix:
      "China (Mainland)_Economic Overview_Apr 28, 2026T19_31 (MONEY & FINANCE - INTEREST RATES).xlsx"
    extracts "interest_rates" → cn_interest_rates.
    """
    base = os.path.splitext(os.path.basename(filename))[0]
    cc = (country or get_country()).lower()

    # China-specific pattern: extract from the trailing parenthetical and
    # strip the all-caps section prefix (MONEY & FINANCE -, INDUSTRY SECTOR -, etc.)
    if base.startswith("China (Mainland)") or base.startswith("Copy of China (Mainland)"):
        # Match the LAST balanced parenthetical, allowing trailing junk
        # (one filename in the dataset has stray ")" after the closing paren).
        matches = list(re.finditer(r"\(([^()]+)\)", base))
        if matches:
            inner = matches[-1].group(1).strip()
            # Drop the all-caps section prefix if present (case-insensitive on
            # the section keyword, since one file uses lowercase 'consumer sector').
            section_prefixes = (
                "MONEY & FINANCE - ", "INDUSTRY SECTOR - ", "SURVEYS & FORECASTS - ",
                "PRICES - ", "EXTERNAL SECTOR - ", "NATIONAL ACCOUNTS - ",
                "LABOR MARKET - ", "GOVERNMENT SECTOR - ", "CONSUMER SECTOR - ",
            )
            for pref in section_prefixes:
                if inner.upper().startswith(pref):
                    inner = inner[len(pref):]
                    break
            body_slug = slugify(inner)
            return f"{cc}_{body_slug}" if body_slug else cc

    # Strip various country-prefix conventions seen in Refinitiv exports
    for prefix in ("US_", "US ", "us_", "us ",
                   "INDONESIA_", "INDONESIA ", "Indonesia_", "Indonesia ",
                   "ID_", "ID ", "id_", "id ",
                   # The Indonesia data also has at least one typo'd file name (Indonsia_)
                   "Indonsia_", "Indonsia "):
        if base.startswith(prefix):
            base = base[len(prefix):]
            break
    body_slug = slugify(base)
    return f"{cc}_{body_slug}" if body_slug else cc


def detect_frequency(dates: list[datetime]) -> str:
    if len(dates) < 2:
        return ""
    deltas_days = []
    for i in range(1, min(20, len(dates))):
        # dates are descending in the sheet (most-recent first); use abs
        d = abs((dates[i - 1] - dates[i]).days)
        if d > 0:
            deltas_days.append(d)
    if not deltas_days:
        return ""
    median = statistics.median(deltas_days)
    if median <= 2:
        return "P1D"
    if median <= 10:
        return "P1W"
    if median <= 45:
        return "P1M"
    if median <= 120:
        return "P3M"
    if median <= 200:
        return "P6M"
    return "P1Y"


def list_xlsx_files(xlsx_dir: str | None = None) -> list[str]:
    base = xlsx_dir if xlsx_dir is not None else get_xlsx_dir()
    files = []
    for name in sorted(os.listdir(base)):
        if name.lower().endswith(".xlsx") and not name.startswith("~$"):
            files.append(os.path.join(base, name))
    return files


def parse_workbook(path: str) -> Iterator[Series]:
    """Yield one Series per RIC in the workbook."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb.active

    section = (ws.cell(4, 1).value or "").strip() if ws.cell(4, 1).value else ""
    category = (ws.cell(4, 2).value or "").strip() if ws.cell(4, 2).value else ""

    source_file = os.path.basename(path)
    cat_slug = slug_from_filename(source_file)
    if not category:
        # Strip any country prefix when synthesizing a human-readable category
        body = re.sub(r"^[a-z]{2}_", "", cat_slug)
        category = body.replace("_", " ").title()

    # Read all rows once into a list (read-only mode is forward-only).
    rows: list[list] = []
    for row in ws.iter_rows(values_only=True):
        rows.append(list(row))
    wb.close()

    if len(rows) < 7:
        return

    header_row = rows[3] if len(rows) > 3 else []  # R4
    ric_row = rows[4] if len(rows) > 4 else []     # R5
    data_rows = rows[6:]                           # R7+

    # RIC columns are at indices 3, 6, 9, 12, ... (zero-based) i.e. 4,7,10,... 1-based
    n_cols = max(len(ric_row), len(header_row))
    for c in range(3, n_cols, 3):
        ric_val = ric_row[c] if c < len(ric_row) else None
        if ric_val is None or str(ric_val).strip() == "":
            continue
        ric = str(ric_val).strip()
        desc = ""
        if c < len(header_row) and header_row[c] is not None:
            desc = str(header_row[c]).strip()

        # Walk data rows for this column-pair
        observations: list[tuple[str, float]] = []
        date_objs: list[datetime] = []
        for r in data_rows:
            if c >= len(r):
                continue
            d = r[c]
            v = r[c + 1] if c + 1 < len(r) else None
            if d is None and v is None:
                continue
            if d is None or v is None:
                continue
            if isinstance(d, datetime):
                date_obj = d
            else:
                # try parse string
                try:
                    date_obj = datetime.fromisoformat(str(d))
                except (TypeError, ValueError):
                    continue
            try:
                value = float(v)
            except (TypeError, ValueError):
                continue
            observations.append((date_obj.date().isoformat(), value))
            date_objs.append(date_obj)

        # Sort ascending by date and dedupe (last write wins for same date)
        seen: dict[str, float] = {}
        for date_str, value in observations:
            seen[date_str] = value
        observations_sorted = sorted(seen.items())

        freq = detect_frequency(sorted(date_objs, reverse=True))

        yield Series(
            ric=ric,
            description=desc,
            frequency=freq,
            section=section,
            category=category,
            category_slug=cat_slug,
            source_file=source_file,
            observations=observations_sorted,
        )


def parse_all(xlsx_dir: str = XLSX_DIR) -> Iterator[Series]:
    for path in list_xlsx_files(xlsx_dir):
        yield from parse_workbook(path)
